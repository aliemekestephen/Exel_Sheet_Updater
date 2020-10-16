import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# function to get file name
filename = input('Please enter file name: ')
docSheet =  input('Please enter sheet number: ')
def process_workbook(filename):
    #getting the excel sheet
    wb = xl.load_workbook(filename)
    sheet = wb[docSheet] # entry value = Sheet1

    # iterating through the sheet to get the cells we need
    for row in range(2, sheet.max_row + 1): #+1 is to get the last count in the range
        cell = sheet.cell(row, 3)
        #print(cell.value) #to confirm the correct cells
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        #print(corrected_price) # to confirm the correction

    # adding a bar chart
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

    print('Update was successful!')
    confirm = input('Do you want to update another file? (Y)es or (N)o: ').lower()
    if confirm == "y":
        filename = input('Please enter file name: ')
        process_workbook(filename)
    else:
        print("This is the end")



process_workbook(filename)
print('Update was successful!')